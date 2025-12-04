from contextlib import contextmanager
from dataclasses import dataclass, field
from pathlib import Path
from typing import ClassVar

from docgen.outputs.abstract import AbstractOutput
from docgen.outputs.descriptor import OutputPathDescriptor
from docgen.utils.path import has_been_modified, hash_path,sanitize_path_part
from docgen.utils.table import to_quarto_markdown
import openpyxl
import pandas as pd
try:
    import xlwings as xw
    from ctypes import windll
except ImportError:
    xw = None
    windll = None
from openpyxl.utils import range_to_tuple
from PIL import ImageGrab
import logging

logger = logging.getLogger(__name__)


class XwCache:
    def __init__(self):
        self.toclose=[]

    @staticmethod
    def is_wb_open(wbpath):
        if xw.apps.count:
            for wb in xw.books:
                if XwCache.tansformwbpath(wb.fullname)==XwCache.tansformwbpath(wbpath):
                    return True
        return False

    @staticmethod
    def tansformwbpath(path):
        return str(path).lower()

    def add(self,wbpath):
        book_key = XwCache.tansformwbpath(wbpath)
        if not XwCache.is_wb_open(book_key):
            self.toclose.append(book_key)

    def close_books(self):
        for wbname in filter(XwCache.is_wb_open, set(self.toclose)):
            wb = xw.Book(wbname)
            app = wb.app
            original_alerts = app.display_alerts
            original_interactive = app.interactive
            
            try:
                app.display_alerts = False
                app.interactive = False
                wb.close()
                logger.info(f"Force closed xlwings workbook: {wbname}")
            finally:
                # Restore original settings
                app.display_alerts = original_alerts
                app.interactive = original_interactive
        self.toclose=[]

  

@dataclass
class CoreExcelOutput:
    """
    Etant donnés un classeur, une feuille et une plage de cellules,
    cette classe permet de créer un output Excel.
    """
    wb_path: Path
    range_name: str

    _wbname: str = field(init=False,default=None)
    _range_boudaries: tuple = field(init=False,default=None)
    _range_str: str = field(init=False,default=None)
    _sheet_name:str = field(init=False,default=None)
    _outname:str = field(init=False,default=None)


    _workbook_cache: ClassVar[dict] = {}
    _xw_cache: ClassVar[XwCache] = XwCache()

    def __post_init__(self):
        self.wb_path = Path(self.wb_path).resolve()
        self._wbname = self.wb_path.stem
        self.wb_path_str = str(self.wb_path)

        self._parse_range(self.range_name)

        self._outname = sanitize_path_part(self._wbname + "_" + self._range_str).lower()

    @classmethod
    def clear_caches(cls):
        cls._xw_cache.close_books()
        for wb in cls._workbook_cache.values():
            try:
                wb.close()
            except Exception as e:
                logger.warning(f"Error closing workbook: {e}")
        cls._workbook_cache.clear()

    @property
    def xw_wb(self):
        bp = str(self.wb_path)
        self._xw_cache.add(bp)
        return xw.Book(bp)

    @property
    def xw_sh(self):
        return self.xw_wb.sheets[self._sheet_name]
    
    @property
    def wb(self):
        """Load workbook using openpyxl"""
        if self.wb_path_str not in self._workbook_cache:
            self._workbook_cache[self.wb_path_str] = openpyxl.load_workbook(self.wb_path)
        return self._workbook_cache[self.wb_path_str]

    @property
    def sh(self):
        """Get worksheet"""
        return self.wb[self._sheet_name]

    def _parse_range(self, range_str):
        """Parse Excel range string to get cell boundaries"""
        if self._range_boudaries is None:
            try:
                self._sheet_name,self._range_boudaries = range_to_tuple(range_str)
                self._range_str=range_str
            except ValueError as e:
                pass
        if self._range_boudaries is None:
            try:
                self._range_boudaries = self._parse_range(f"{self.wb.sheetnames[0]}!{range_str}")
                if len(self.wb.sheetnames)>1:
                    logger.warning(f"Aucune feuille définie, {self._sheet_name} retenue")
            except ValueError as e:
                pass
        if self._range_boudaries is None:
            _ = self.sh.defined_names.get(range_str)
            if _ is None:
                _ = self.wb.defined_names.get(range_str)
            if _ is not None:
                g = iter(_.destinations)
                sheet_name,range_ = next(g)
                self._range_boudaries = self._parse_range(f"{sheet_name}!{range_}")
                if next(g,None) is not None:
                    raise ValueError("Les plages multidestinations ne sont pas supportées")
        if self._range_boudaries is None:
            raise ValueError(f"Invalid range: {range_str}")
        return self._range_boudaries

    def _get_range_values(self)->pd.DataFrame:
        """Extract values from Excel range"""
        min_col, min_row, max_col, max_row = self._range_boudaries
        values = []
        for row in self.sh.iter_rows(min_row=min_row, max_row=max_row, 
                                     min_col=min_col, max_col=max_col, values_only=True):
            values.append(list(row))
        return pd.DataFrame(values)
    
    def build(self,mode="paragraph",parent_path:Path=None,**kwargs)->Path:
        """
        Build the output based on the specified mode only if the file has been modified.
        """
        if parent_path is None:
            parent_path = self.wb_path.parent
        path = (parent_path / self._outname)
        if mode == "paragraph":
            path = path.with_suffix(".txt")
            f= self.build_paragraph
        elif mode == "image":
            path = path.with_suffix(".png")
            f = self.build_image
        elif mode == "html":
            path = path.with_suffix(".html")
            f = self.build_html
        elif mode in ["markdown","quarto_markdown"]:
            path = path.with_suffix(".md")
            f = self.build_markdown if mode == "markdown" else self.build_quarto_markdown
        else:
            raise ValueError(f"Mode {mode} non supporté. Utiliser 'paragraph', 'image' ou 'html'.")
        if has_been_modified(self.wb_path, path, hash=False):
            f(path,**kwargs)
            logger.info(f"Output créé: {path}")
        return path

    def build_paragraph(self,path:Path):
        _ = self._get_range_values()
        _ = _.fillna('').astype(str).agg(" ".join, axis=1)
        _ = _.str.replace(r'\s+', ' ', regex=True)
        _ = _.str.strip()
        _ = _.str.cat(sep="\n")
        # _ = _[_ != '']  # Supprimer les lignes vides
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(_, encoding="utf8")
    
    def build_image(self,path:Path):
        if windll is None or xw is None:
            raise RuntimeError("xlwings and ctypes.windll are required for image output.")
        self.xw_sh.range(self.range_name).api.Copy()
        windll.user32.OpenClipboard(None)
        img = ImageGrab.grabclipboard()
        img.save(path)
        if windll.user32.OpenClipboard(None):
            windll.user32.EmptyClipboard()
            windll.user32.CloseClipboard()
    
    def build_html(self,path:Path,**kwargs):
        plage = self._get_range_values()
        kwargs.setdefault("header", False)
        kwargs.setdefault("index", False)
        kwargs.setdefault("table_id", self._outname)
        plage.to_html(path,**kwargs)

    def build_quarto_markdown(self, path: Path, **kwargs):
        """
        Build the output as Quarto Markdown.
        """
        kwargs.setdefault("table_id", self._outname)
        markdown_content = to_quarto_markdown(self._get_range_values(), **kwargs)
        path.parent.mkdir(parents=True, exist_ok=True)
        path.write_text(markdown_content, encoding="utf8")
    
    def build_markdown(self,path:Path,**kwargs):
        plage = self._get_range_values()
        kwargs.setdefault("index", self.with_rows)
        kwargs.setdefault("tablefmt","grid")
        plage.to_markdown(path,**kwargs)

@dataclass
class ExcelOutput(AbstractOutput):
    
    @classmethod
    @contextmanager
    def executor(cls):
        try:
            yield
        finally:
            CoreExcelOutput.clear_caches()

    def build(self):
        label = self.rematch.group(1).strip()
        excel_path = self.rematch.group(2).strip()
        descr = OutputPathDescriptor.from_string(excel_path,
                                                 source_dir=self.container.source_dir,
                                                 dest_dir=self.container.build_dir,)
        excel_range = self.rematch.group(3).strip()
        self.dest = self.generated_dir / hash_path(descr.absolute_path,n_hash=8,n_parents=2)
        self.dest.mkdir(parents=True, exist_ok=True)
        self.label = label
        self.core_out = CoreExcelOutput(
            descr.absolute_path,
            range_name=excel_range,
        )


class ExcelImgOutput(ExcelOutput):
    """
    Output pour les images Excel.
    """
    def build(self):
        super().build()
        result = self.core_out.build(mode="image", parent_path=self.dest)
        result = result.relative_to(self.container.build_dir)
        self.sub_by = f"![{self.label}]({str(result)})"
    
class ExcelMarkdownOutput(ExcelOutput):
    """
    Output pour les tableaux Excel.
    """
    def build(self):
        super().build()
        # Approche Markdown
#         result = self.core_out.build(mode="markdown", parent_path=self.dest)
#         content = result.read_text(encoding="utf8")
#         result.write_text(f"""
# {content}
# : {self.label} {{.striped .hover}}
# """,encoding="utf8")

        # Approche quarto_markdown (via html)
        result = self.core_out.build(mode="quarto_markdown", parent_path=self.dest)
        
        # Common
        result = result.relative_to(self.container.build_dir)
        self.sub_by = "{{"+f"< include {str(result)} >"+"}}"


if __name__ == "__main__":
    pass